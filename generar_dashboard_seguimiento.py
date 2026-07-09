#!/usr/bin/env python3
"""
Genera/actualiza el dashboard "Sistema de Seguimiento OT — Producción"
a partir del archivo Excel de seguimiento (SEGUIMIENTO_DE_OT1.xlsx o el
nombre que corresponda) y una plantilla HTML existente.

USO:
    python3 generar_dashboard_seguimiento.py \
        --excel "SEGUIMIENTO_DE_OT1.xlsx" \
        --plantilla "Sistema_de_Seguimiento_OT___Produccion_Fix_Filter.htm" \
        --salida "Sistema_de_Seguimiento_OT___Produccion_Fix_Filter.htm"

Si no se pasan argumentos, usa los valores por defecto definidos abajo en
CONFIG, pensados para correr dentro del workflow de GitHub Actions.
"""

import argparse
import json
import re
import sys
import time
import random
import datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

EXCEL_EPOCH = datetime.datetime(1899, 12, 30)

# Cada hoja del Excel corresponde a un área de producción (proc).
# 'col_kg', 'col_ml', 'col_piezas', 'col_rollos', 'col_desp' son los
# encabezados EXACTOS tal como aparecen en la fila 1 de cada hoja.
# Si el encabezado no aplica para esa hoja, se deja en None y el campo
# se guarda como 0.
SHEETS = {
    "IMPRESION": {
        "proc": "IMP",
        "col_kg": "KILOS IMPRESOS",
        "col_ml": "METROS LINEALES IMPRESOS",
        "col_piezas": None,
        "col_rollos": "No. DE ROLLOS IMPRESOS",
        "col_desp": "DESPERDICIO REPORTADO",
    },
    "LAMINACION": {
        "proc": "LAM",
        "col_kg": "KG LAMINADOS",
        "col_ml": "METROS LINEALES LAMINADOS",
        "col_piezas": None,
        "col_rollos": None,
        "col_desp": "DESPERDICIO REPORTADO",
    },
    "REFINADO": {
        "proc": "REF",
        "col_kg": "KG REFINADOS",
        "col_ml": "ML REFINADOS",
        "col_piezas": None,
        "col_rollos": None,
        "col_desp": "DESPERDICIO REPORTADO",
    },
    "POUCH": {
        "proc": "POUCH",
        "col_kg": "KILOGRAMOS",
        "col_ml": None,
        "col_piezas": "PIEZAS",
        "col_rollos": None,
        "col_desp": "DESPERDICIO REPORTADO",
    },
    "SIDE WELD": {
        "proc": "SW",
        "col_kg": "KILOGRAMOS",
        "col_ml": None,
        "col_piezas": "PIEZAS",
        "col_rollos": None,
        "col_desp": "DESPERDICIO REPORTADO",
    },
}

# Columnas comunes a todas las hojas (encabezados exactos).
COL_ORDEN = "ORDEN"
COL_FECHA = "FECHA"
COL_MAQUINA = "MAQUINA"
COL_TURNO = "TURNO"          # se normaliza quitando espacios/mayúsculas al buscar
COL_OPERADOR = "OPERADOR"
COL_SUPERVISOR = "SUPERVISOR"

DEFAULT_EXCEL = "SEGUIMIENTO_DE_OT1.xlsx"
DEFAULT_HTML = "index.html"

# Hoja y columnas usadas para construir el "catalogo" (info de la OT:
# cliente, descripción, kg especificados). Se usa coincidencia por
# substring en la columna de descripción porque el encabezado original
# viene con problemas de codificación ("DescripciÃ³n").
CATALOGO_SHEET = "rep-ov-espec"
CATALOGO_COL_OT = "ID Orden de Venta"
CATALOGO_COL_CLIENTE = "Nombre Cliente"
CATALOGO_COL_DESCRIPCION_CONTAINS = "ESCRIPCI"   # matchea "Descripción" con o sin mojibake
CATALOGO_COL_KGESPEC = "Cantidad Kilos"

DATA_BLOCK_RE = re.compile(
    r'(<script id="datos-incrustados" type="application/json">)(.*?)(</script>)',
    re.DOTALL,
)


# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------

def to_iso_date(value):
    """Convierte un valor de celda FECHA (datetime, número serial de Excel
    o texto) al formato de texto 'YYYY-MM-DD' que usa la versión actual
    del dashboard."""
    if value is None or value == "":
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
            value = datetime.datetime(value.year, value.month, value.day)
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        # número serial de Excel
        return (EXCEL_EPOCH + datetime.timedelta(days=value)).strftime("%Y-%m-%d")
    # último intento: texto que ya viene como fecha
    try:
        parsed = datetime.datetime.fromisoformat(str(value).strip())
        return parsed.strftime("%Y-%m-%d")
    except Exception:
        return str(value).strip() or None


def to_number(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return 0


def build_header_index(header_row):
    """Devuelve dict {nombre_columna_normalizado: índice} para una fila de encabezados."""
    idx = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().upper()
        idx[key] = i
    return idx


def find_col(header_idx, name):
    if name is None:
        return None
    key = name.strip().upper()
    if key in header_idx:
        return header_idx[key]
    # búsqueda flexible por si trae espacios extra (ej. "TURNO " vs "TURNO")
    for k, i in header_idx.items():
        if k.replace(" ", "") == key.replace(" ", ""):
            return i
    return None


def gen_id(base_ms, i):
    # Réplica del patrón usado por el dashboard original: timestamp en ms
    # + fracción para diferenciar filas generadas en el mismo lote.
    return round(base_ms + i, 4) + round(random.random(), 4)


# ---------------------------------------------------------------------------
# LECTURA DEL EXCEL
# ---------------------------------------------------------------------------

def leer_registros(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    registros = []
    base_ms = int(time.time() * 1000)
    counter = 0

    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [aviso] Hoja '{sheet_name}' no encontrada en el Excel, se omite.")
            continue

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            continue
        header_idx = build_header_index(header_row)

        c_orden = find_col(header_idx, COL_ORDEN)
        c_fecha = find_col(header_idx, COL_FECHA)
        c_maquina = find_col(header_idx, COL_MAQUINA)
        c_turno = find_col(header_idx, COL_TURNO)
        c_operador = find_col(header_idx, COL_OPERADOR)
        c_supervisor = find_col(header_idx, COL_SUPERVISOR)
        c_kg = find_col(header_idx, cfg["col_kg"])
        c_ml = find_col(header_idx, cfg["col_ml"])
        c_piezas = find_col(header_idx, cfg["col_piezas"])
        c_rollos = find_col(header_idx, cfg["col_rollos"])
        c_desp = find_col(header_idx, cfg["col_desp"])

        if c_orden is None:
            print(f"  [aviso] Hoja '{sheet_name}': no se encontró columna ORDEN, se omite.")
            continue

        n_hoja = 0
        for row in rows_iter:
            if row is None:
                continue
            orden = row[c_orden] if c_orden < len(row) else None
            if orden is None or str(orden).strip() == "":
                continue

            fecha_val = row[c_fecha] if c_fecha is not None and c_fecha < len(row) else None
            registro = {
                "id": gen_id(base_ms, counter),
                "ot": str(orden).strip(),
                "proc": cfg["proc"],
                "origen": "excel",
                "fecha": to_iso_date(fecha_val),
                "maquina": str(row[c_maquina]).strip() if c_maquina is not None and c_maquina < len(row) and row[c_maquina] is not None else "",
                "turno": str(row[c_turno]).strip() if c_turno is not None and c_turno < len(row) and row[c_turno] is not None else "",
                "operador": str(row[c_operador]).strip() if c_operador is not None and c_operador < len(row) and row[c_operador] is not None else "",
                "supervisor": str(row[c_supervisor]).strip() if c_supervisor is not None and c_supervisor < len(row) and row[c_supervisor] is not None else "",
                "kg": to_number(row[c_kg]) if c_kg is not None and c_kg < len(row) else 0,
                "ml": to_number(row[c_ml]) if c_ml is not None and c_ml < len(row) else 0,
                "piezas": to_number(row[c_piezas]) if c_piezas is not None and c_piezas < len(row) else 0,
                "desp": to_number(row[c_desp]) if c_desp is not None and c_desp < len(row) else 0,
                "rollos": to_number(row[c_rollos]) if c_rollos is not None and c_rollos < len(row) else 0,
            }
            registros.append(registro)
            counter += 1
            n_hoja += 1

        print(f"  Hoja '{sheet_name}' ({cfg['proc']}): {n_hoja} registros")

    wb.close()
    return registros


def construir_ots(registros):
    ots = {}
    for r in registros:
        ot = r["ot"]
        proc = r["proc"]
        fecha = r["fecha"] or ""   # las fechas ISO "YYYY-MM-DD" se comparan bien como texto
        entry = ots.setdefault(ot, {})
        acc = entry.setdefault(proc, {"kg": 0, "ml": 0, "piezas": 0, "desp": 0, "n": 0, "ult": ""})
        acc["kg"] += r["kg"]
        acc["ml"] += r["ml"]
        acc["piezas"] += r["piezas"]
        acc["desp"] += r["desp"]
        acc["n"] += 1
        if fecha > acc["ult"]:
            acc["ult"] = fecha
    return ots


def construir_catalogo(excel_path):
    """Lee la hoja 'rep-ov-espec' y arma {ot: {cliente, descripcion, kgEspec}}.
    Si una OT aparece en varias filas (varias líneas de especificación),
    se conserva la última encontrada."""
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if CATALOGO_SHEET not in wb.sheetnames:
        print(f"  [aviso] Hoja '{CATALOGO_SHEET}' no encontrada, catalogo quedará vacío.")
        wb.close()
        return {}

    ws = wb[CATALOGO_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return {}

    header_idx = build_header_index(header_row)
    c_ot = find_col(header_idx, CATALOGO_COL_OT)
    c_cliente = find_col(header_idx, CATALOGO_COL_CLIENTE)
    c_kg = find_col(header_idx, CATALOGO_COL_KGESPEC)

    c_desc = None
    for k, i in header_idx.items():
        if CATALOGO_COL_DESCRIPCION_CONTAINS in k:
            c_desc = i
            break

    if c_ot is None:
        print(f"  [aviso] No se encontró columna '{CATALOGO_COL_OT}' en '{CATALOGO_SHEET}'.")
        wb.close()
        return {}

    catalogo = {}
    n = 0
    for row in rows_iter:
        if row is None:
            continue
        ot = row[c_ot] if c_ot < len(row) else None
        if ot is None or str(ot).strip() == "":
            continue
        ot = str(ot).strip()
        catalogo[ot] = {
            "cliente": str(row[c_cliente]).strip() if c_cliente is not None and c_cliente < len(row) and row[c_cliente] is not None else "",
            "descripcion": str(row[c_desc]).strip() if c_desc is not None and c_desc < len(row) and row[c_desc] is not None else "",
            "kgEspec": to_number(row[c_kg]) if c_kg is not None and c_kg < len(row) else 0,
        }
        n += 1

    wb.close()
    print(f"  Catálogo: {n} filas leídas de '{CATALOGO_SHEET}', {len(catalogo)} OT únicas")
    return catalogo


# ---------------------------------------------------------------------------
# INYECCIÓN EN EL HTML
# ---------------------------------------------------------------------------

def actualizar_html(plantilla_path, salida_path, data_dict):
    html = Path(plantilla_path).read_text(encoding="utf-8")

    if not DATA_BLOCK_RE.search(html):
        raise RuntimeError(
            "No se encontró el bloque <script id=\"datos-incrustados\" "
            "type=\"application/json\">...</script> en la plantilla. "
            "Verifica que 'plantilla' sea el HTML correcto."
        )

    nuevo_json = json.dumps(data_dict, ensure_ascii=False, separators=(",", ":"))

    def _replace(match):
        return match.group(1) + nuevo_json + match.group(3)

    html_actualizado = DATA_BLOCK_RE.sub(_replace, html, count=1)
    Path(salida_path).write_text(html_actualizado, encoding="utf-8")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Actualiza el dashboard de Seguimiento de OT")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Ruta al archivo Excel de origen")
    parser.add_argument("--plantilla", default=DEFAULT_HTML, help="Ruta al HTML existente (se usa como plantilla)")
    parser.add_argument("--salida", default=DEFAULT_HTML, help="Ruta del HTML de salida (puede ser el mismo que la plantilla)")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    plantilla_path = Path(args.plantilla)

    if not excel_path.exists():
        print(f"ERROR: no se encontró el Excel: {excel_path}", file=sys.stderr)
        sys.exit(1)
    if not plantilla_path.exists():
        print(f"ERROR: no se encontró la plantilla HTML: {plantilla_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Leyendo Excel: {excel_path}")
    registros = leer_registros(excel_path)
    print(f"Total de registros leídos: {len(registros)}")

    print("Construyendo agregación por OT...")
    ots = construir_ots(registros)
    print(f"Total de OT únicas: {len(ots)}")

    print("Construyendo catálogo (cliente / descripción / kg especificados)...")
    catalogo = construir_catalogo(excel_path)

    data_dict = {"registros": registros, "ots": ots, "catalogo": catalogo}

    print(f"Actualizando HTML -> {args.salida}")
    actualizar_html(plantilla_path, args.salida, data_dict)

    print("Listo.")


if __name__ == "__main__":
    main()
